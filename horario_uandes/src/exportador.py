"""
exportador.py — Módulo 4: Exportación a Excel.

Genera el archivo de salida con dos hojas:
  - HORARIO: listado completo de secciones asignadas (una fila por sección).
  - REPORTE: métricas generales y resumen de violaciones de restricciones.

Uso:
    from exportador import exportar_excel
    exportar_excel(datos, "outputs/horario_generado.xlsx")
"""

from collections import defaultdict
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import DatosProblema, TipoReunion, TipoProfesor

# Colores por tipo de reunión (ARGB sin alpha para openpyxl)
_COLOR_TIPO: dict[TipoReunion, str] = {
    TipoReunion.CLASE:       "FFD9EAD3",  # verde claro
    TipoReunion.AYUDANTIA:   "FFFCE5CD",  # naranja claro
    TipoReunion.LABORATORIO: "FFD0E4F7",  # azul claro
    TipoReunion.PRUEBA:      "FFFFF2CC",  # amarillo claro
    TipoReunion.EXAMEN:      "FFFFD6CC",  # rojo claro
}

_DIAS_ORDEN = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes"]


def exportar_excel(datos: DatosProblema, ruta_salida: str) -> None:
    """Genera el Excel de salida con hojas HORARIO y REPORTE.

    Args:
        datos: DatosProblema con bloque_asignado rellenado por los módulos
               CP-SAT y GA.
        ruta_salida: Ruta del archivo de salida (ej: "outputs/horario.xlsx").
    """
    print("=" * 60)
    print("MÓDULO 4: Exportador Excel")
    print("=" * 60)

    Path(ruta_salida).parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # eliminar hoja vacía por defecto

    _escribir_horario(wb, datos)
    _escribir_reporte(wb, datos)

    wb.save(ruta_salida)
    print(f"\n  Archivo guardado: {ruta_salida}")


# =============================================================================
# HOJA HORARIO
# =============================================================================

def _escribir_horario(wb: openpyxl.Workbook, datos: DatosProblema) -> None:
    """Hoja HORARIO: una fila por sección asignada, ordenada por día y hora."""
    ws = wb.create_sheet("HORARIO")

    encabezados = [
        "NRC", "Código", "Título", "Sección",
        "Tipo", "Profesor", "Día", "Hora Inicio", "Hora Fin",
        "Sala Especial", "Cupos",
    ]

    # Estilo de encabezado
    header_fill = PatternFill(fill_type="solid", fgColor="FF4472C4")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, texto in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col, value=texto)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 30

    # Construir filas
    filas = []
    for s in datos.secciones:
        if s.bloque_asignado is None:
            continue
        b = s.bloque_asignado
        filas.append({
            "NRC": s.nrc,
            "Código": s.curso.codigo,
            "Título": s.curso.titulo,
            "Sección": s.numero_seccion,
            "Tipo": s.tipo_reunion.value,
            "Profesor": s.profesor.nombre if s.profesor else "",
            "Día": b.dia.value,
            "Hora Inicio": b.hora_inicio,
            "Hora Fin": b.hora_fin,
            "Sala Especial": s.sala_especial.nombre_corto if s.sala_especial else "",
            "Cupos": s.cupos,
        })

    # Ordenar: día → hora_inicio → código → sección
    orden_dia = {d: i for i, d in enumerate(_DIAS_ORDEN)}
    filas.sort(key=lambda f: (
        orden_dia.get(f["Día"], 99),
        f["Hora Inicio"],
        f["Código"],
        f["Sección"],
    ))

    # Escribir filas
    data_align = Alignment(vertical="center")
    for fila_idx, fila in enumerate(filas, start=2):
        tipo_enum = TipoReunion(fila["Tipo"])
        fill_color = PatternFill(fill_type="solid", fgColor=_COLOR_TIPO.get(tipo_enum, "FFFFFFFF"))

        for col, key in enumerate(encabezados, start=1):
            cell = ws.cell(row=fila_idx, column=col, value=fila[key])
            cell.fill = fill_color
            cell.alignment = data_align
            cell.border = border

    # Anchos de columna
    anchos = [10, 10, 40, 8, 6, 35, 12, 12, 12, 30, 8]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # Congelar primera fila
    ws.freeze_panes = "A2"

    # Autofilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(encabezados))}1"

    asignadas = len(filas)
    print(f"\n  Hoja HORARIO: {asignadas} secciones asignadas")


# =============================================================================
# HOJA REPORTE
# =============================================================================

def _escribir_reporte(wb: openpyxl.Workbook, datos: DatosProblema) -> None:
    """Hoja REPORTE: métricas de calidad y secciones especiales."""
    ws = wb.create_sheet("REPORTE")

    titulo_font = Font(bold=True, size=13)
    seccion_font = Font(bold=True, size=11, color="FF4472C4")
    bold = Font(bold=True)
    fill_seccion = PatternFill(fill_type="solid", fgColor="FFD9E1F2")
    thin = Side(style="thin", color="FFBFBFBF")
    border_light = Border(bottom=thin)

    fila = 1

    def _titulo(texto: str) -> None:
        nonlocal fila
        cell = ws.cell(row=fila, column=1, value=texto)
        cell.font = titulo_font
        fila += 1

    def _seccion(texto: str) -> None:
        nonlocal fila
        fila += 1
        cell = ws.cell(row=fila, column=1, value=texto)
        cell.font = seccion_font
        cell.fill = fill_seccion
        ws.merge_cells(f"A{fila}:D{fila}")
        fila += 1

    def _fila_dato(clave: str, valor) -> None:
        nonlocal fila
        ws.cell(row=fila, column=1, value=clave).font = bold
        ws.cell(row=fila, column=2, value=valor)
        fila += 1

    def _encabezados_tabla(cols: list) -> None:
        nonlocal fila
        for col, texto in enumerate(cols, start=1):
            c = ws.cell(row=fila, column=col, value=texto)
            c.font = bold
            c.border = border_light
        fila += 1

    # -------------------------------------------------------------------------
    # 1. Resumen general
    # -------------------------------------------------------------------------
    _titulo("REPORTE DE GENERACIÓN DE HORARIO")
    _fila_dato("Fecha de generación", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"))
    fila += 1

    total = len(datos.secciones)
    asignadas = sum(1 for s in datos.secciones if s.bloque_asignado is not None)
    omitidas_2mas1 = sum(
        1 for s in datos.secciones if s.bloque_asignado is None
        and ("2+1" in s.curso.distribucion or "2-1" in s.curso.distribucion)
    )
    sin_asignar = total - asignadas - omitidas_2mas1

    _seccion("1. RESUMEN GENERAL")
    _fila_dato("Total de secciones en el maestro", total)
    _fila_dato("Secciones asignadas (con bloque horario)", asignadas)
    _fila_dato("Secciones omitidas (distribución 2+1)", omitidas_2mas1)
    _fila_dato("Secciones sin asignar (error)", sin_asignar)
    _fila_dato("Tasa de asignación", f"{asignadas / total * 100:.1f}%")

    # Conteo por tipo
    _seccion("2. DESGLOSE POR TIPO DE REUNIÓN")
    conteo_tipo: dict = defaultdict(lambda: {"total": 0, "asignadas": 0})
    for s in datos.secciones:
        tipo = s.tipo_reunion.value
        conteo_tipo[tipo]["total"] += 1
        if s.bloque_asignado:
            conteo_tipo[tipo]["asignadas"] += 1
    _encabezados_tabla(["Tipo", "Total", "Asignadas", "Omitidas"])
    for tipo, cnt in sorted(conteo_tipo.items()):
        ws.cell(row=fila, column=1, value=tipo)
        ws.cell(row=fila, column=2, value=cnt["total"])
        ws.cell(row=fila, column=3, value=cnt["asignadas"])
        ws.cell(row=fila, column=4, value=cnt["total"] - cnt["asignadas"])
        fila += 1

    # -------------------------------------------------------------------------
    # 2. Distribución por día y bloque horario
    # -------------------------------------------------------------------------
    _seccion("3. DISTRIBUCIÓN DE SECCIONES ASIGNADAS POR DÍA")
    por_dia: dict = defaultdict(int)
    for s in datos.secciones:
        if s.bloque_asignado:
            por_dia[s.bloque_asignado.dia.value] += 1
    _encabezados_tabla(["Día", "Secciones"])
    for dia in _DIAS_ORDEN:
        ws.cell(row=fila, column=1, value=dia)
        ws.cell(row=fila, column=2, value=por_dia.get(dia, 0))
        fila += 1

    # -------------------------------------------------------------------------
    # 3. Restricciones blandas
    # -------------------------------------------------------------------------
    _seccion("4. VIOLACIONES DE RESTRICCIONES BLANDAS")

    # RB2: profesores JORNADA en extremos
    _EXTREMOS_MIN = {8 * 60 + 30, 17 * 60 + 30}

    def _hora_a_min(hora: str) -> int:
        h, m = hora.split(":")
        return int(h) * 60 + int(m)

    rb2_viols = [
        s for s in datos.secciones
        if s.bloque_asignado and s.profesor
        and s.profesor.tipo == TipoProfesor.JORNADA
        and _hora_a_min(s.bloque_asignado.hora_inicio) in _EXTREMOS_MIN
    ]
    _fila_dato("RB2 - Prof. jornada en bloque extremo (8:30 ó 17:30)", len(rb2_viols))
    if rb2_viols:
        _encabezados_tabla(["Sección", "Profesor", "Día", "Hora"])
        for s in rb2_viols[:20]:
            b = s.bloque_asignado
            ws.cell(row=fila, column=1, value=s.id)
            ws.cell(row=fila, column=2, value=s.profesor.nombre)
            ws.cell(row=fila, column=3, value=b.dia.value)
            ws.cell(row=fila, column=4, value=b.hora_inicio)
            fila += 1
        if len(rb2_viols) > 20:
            ws.cell(row=fila, column=1, value=f"... y {len(rb2_viols) - 20} más")
            fila += 1

    # RB3/RB4: pares del mismo NRC en mismo día
    por_nrc: dict = defaultdict(list)
    for s in datos.secciones:
        if s.bloque_asignado:
            por_nrc[s.nrc].append(s)

    rb3_count = 0
    rb3_ejemplos = []
    for nrc, secs in por_nrc.items():
        if len(secs) < 2:
            continue
        dias = [s.bloque_asignado.dia for s in secs]
        for a in range(len(secs)):
            for b in range(a + 1, len(secs)):
                if dias[a] == dias[b]:
                    rb3_count += 1
                    if len(rb3_ejemplos) < 10:
                        rb3_ejemplos.append((secs[a], secs[b]))

    fila += 1
    _fila_dato("RB3 - Pares de sección misma NRC en mismo día", rb3_count)
    if rb3_ejemplos:
        _encabezados_tabla(["Sección A", "Sección B", "Día", "Horas"])
        for s1, s2 in rb3_ejemplos:
            b1, b2 = s1.bloque_asignado, s2.bloque_asignado
            ws.cell(row=fila, column=1, value=s1.id)
            ws.cell(row=fila, column=2, value=s2.id)
            ws.cell(row=fila, column=3, value=b1.dia.value)
            ws.cell(row=fila, column=4, value=f"{b1.hora_inicio} / {b2.hora_inicio}")
            fila += 1

    # -------------------------------------------------------------------------
    # 4. Secciones omitidas (2+1)
    # -------------------------------------------------------------------------
    omitidas = [s for s in datos.secciones if s.bloque_asignado is None]
    _seccion(f"5. SECCIONES NO ASIGNADAS ({len(omitidas)} total)")
    if omitidas:
        _encabezados_tabla(["ID Sección", "Código", "Título", "Distribución", "Razón"])
        for s in omitidas:
            razon = "Distribución 2+1 (fuera de scope v1)" if (
                "2+1" in s.curso.distribucion or "2-1" in s.curso.distribucion
            ) else "Dominio vacío o restricción incompatible"
            ws.cell(row=fila, column=1, value=s.id)
            ws.cell(row=fila, column=2, value=s.curso.codigo)
            ws.cell(row=fila, column=3, value=s.curso.titulo)
            ws.cell(row=fila, column=4, value=s.curso.distribucion)
            ws.cell(row=fila, column=5, value=razon)
            fila += 1

    # Anchos de columnas
    anchos_rep = [40, 20, 20, 25, 50]
    for col, ancho in enumerate(anchos_rep, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    print(f"  Hoja REPORTE: generada con métricas y {len(omitidas)} secciones no asignadas")
